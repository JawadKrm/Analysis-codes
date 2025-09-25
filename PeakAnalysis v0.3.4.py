"""
PeakAnalysis v0.3.4 (Prototype)

Changes from v0.3.3:
- Accepts .smr or .npy. If .npy is not dictionary-shaped the GUI can convert it automatically.
- Default FFT preference changed to 16.
- Task checkbox in setup: user can specify task start/end; saved peaks will note if window overlaps the task.
- Save-peak popup now has "Don't show this again" checkbox to silence future popups.
- Export into existing spreadsheet supports several modes:
    - All files into selected/new sheet
    - Each file in separate sheets
    - Export separate files for each analysis
- Many UX issues fixed per specification.
Version: v0.3.4
Author: Jawad_Karim (refactor)
Date: 2025-09-24
"""
import os
import math
import csv
import traceback
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from scipy.signal import welch

try:
    import neo
except Exception:
    neo = None

# Optional XLSX support
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
except Exception:
    openpyxl = None

# -------------------------
# Utilities (conversion + loaders + peak finding)
# -------------------------
def list_power_of_two(min_pow=4, max_pow=18):
    return [2 ** p for p in range(min_pow, max_pow + 1)]

def get_sampling_rate_hz(signal):
    """Robust extraction of sampling rate from neo analogsignal-like object."""
    sr = getattr(signal, "sampling_rate", None)
    if sr is None:
        raise ValueError("Signal has no sampling_rate attribute")
    try:
        return float(sr.rescale("Hz").magnitude)
    except Exception:
        try:
            return float(sr)
        except Exception:
            raise ValueError("Unable to determine sampling rate (Hz) from signal.")

def parabolic_peak_interpolation(freqs, powers, peak_idx, use_log=True, eps=None):
    freqs = np.asarray(freqs, dtype=float)
    powers = np.asarray(powers, dtype=float)
    N = len(powers)
    if N == 0:
        return float('nan'), float('nan')
    if peak_idx <= 0 or peak_idx >= N - 1:
        return float(freqs[peak_idx]), float(max(0.0, powers[peak_idx]))
    x = freqs[peak_idx - 1:peak_idx + 2].astype(float)
    y = powers[peak_idx - 1:peak_idx + 2].astype(float)
    if eps is None:
        eps = max(np.max(y), 1.0) * 1e-12 + 1e-20
    if use_log:
        y = np.log(y + eps)
    dxs = np.diff(x)
    if np.allclose(dxs, dxs[0], rtol=1e-6, atol=1e-12):
        dx = dxs[0]
        y0, y1, y2 = y
        denom = (y0 - 2.0 * y1 + y2)
        if denom == 0.0:
            delta = 0.0
        else:
            delta = 0.5 * (y0 - y2) / denom
        peak_freq = x[1] + delta * dx
        if use_log:
            peak_power = float(np.exp(y1 - 0.25 * (y0 - y2) * delta) - eps)
        else:
            peak_power = float(y1 - 0.25 * (y0 - y2) * delta)
    else:
        a, b, c = np.polyfit(x, y, 2)
        if np.isclose(a, 0.0):
            peak_freq = float(x[1])
            peak_power = float(np.exp(y[1]) - eps) if use_log else float(y[1])
        else:
            peak_freq = float(-b / (2.0 * a))
            y_at_peak = np.polyval([a, b, c], peak_freq)
            peak_power = float(np.exp(y_at_peak) - eps) if use_log else float(y_at_peak)
    if peak_power < 0:
        peak_power = 0.0
    return peak_freq, peak_power

def find_beta_peak(freqs, pxx_bin, beta_band=(13.0, 30.0)):
    mask = (freqs >= beta_band[0]) & (freqs <= beta_band[1])
    if not np.any(mask):
        return float('nan'), float('nan')
    f_beta = freqs[mask]
    p_beta = pxx_bin[mask]
    if len(p_beta) == 0:
        return float('nan'), float('nan')
    rel_peak_idx = int(np.argmax(p_beta))
    return parabolic_peak_interpolation(f_beta, p_beta, rel_peak_idx, use_log=True)

# Robust npy loader / converter used by GUI
def inspect_npy_file(path):
    """
    Inspect .npy file, return a tuple (is_dict, info_string, loaded_obj)
    is_dict True if already dictionary-like (object array with dict or direct dict)
    """
    arr = np.load(path, allow_pickle=True)
    # If 0-d object array wrapping a dict
    if hasattr(arr, 'shape') and arr.shape == ():
        try:
            obj = arr.item()
            if isinstance(obj, dict):
                return True, f"0-d object array wrapping dict with keys: {list(obj.keys())}", obj
        except Exception:
            pass
    # If arr is dict itself
    if isinstance(arr, dict):
        return True, f"Loaded dict with keys: {list(arr.keys())}", arr
    # Otherwise return raw array and info
    return False, f"Loaded type={type(arr)}, shape={getattr(arr,'shape',None)}, dtype={getattr(arr,'dtype',None)}", arr

def force_convert_npy_to_dict(arr, default_fs=125.0):
    """
    Try to convert arbitrary loaded .npy object (arr) into dictionary:
      { 'ch1': {'sampling_rate': fs, 'data': np.ndarray}, ... }
    This mirrors the repair script logic.
    Returns dict.
    """
    # If arr is 0-d dict-like
    if isinstance(arr, dict):
        out = {}
        idx = 1
        for k, v in arr.items():
            if isinstance(v, dict) and 'data' in v:
                data = np.asarray(v['data'], dtype=float)
                fs = v.get('sampling_rate', default_fs)
            else:
                data = np.asarray(v)
                fs = default_fs
            data_clean = _clean_channel_array(data)
            out[f"ch{idx}"] = {'sampling_rate': float(fs) if fs is not None else None, 'data': data_clean, 'label': str(k)}
            idx += 1
        return out

    a = np.asarray(arr, dtype=object) if not isinstance(arr, np.ndarray) else arr
    # if object array of length 1 and element is dict
    if hasattr(a, 'shape') and a.shape == () and isinstance(a.item(), dict):
        return force_convert_npy_to_dict(a.item(), default_fs)

    # if numeric ndarray
    if isinstance(a, np.ndarray):
        if a.ndim == 1:
            # possibly object array with elements each channel
            if a.dtype == object:
                out = {}
                idx = 1
                for el in a:
                    if isinstance(el, dict) and 'data' in el:
                        d = np.asarray(el['data'])
                        fs = el.get('sampling_rate', default_fs)
                    else:
                        d = np.asarray(el)
                        fs = default_fs
                    out[f"ch{idx}"] = {'sampling_rate': float(fs) if fs is not None else None, 'data': _clean_channel_array(d)}
                    idx += 1
                return out
            else:
                # numeric 1D -> single channel
                return {'ch1': {'sampling_rate': default_fs, 'data': _clean_channel_array(a)}}
        elif a.ndim == 2:
            # heuristic: rows=time, cols=channels if rows >= cols
            r, c = a.shape
            out = {}
            if r >= c:
                # treat columns as channels
                for i in range(c):
                    col = a[:, i]
                    out[f"ch{i+1}"] = {'sampling_rate': default_fs, 'data': _clean_channel_array(col)}
            else:
                # transpose
                at = a.T
                for i in range(at.shape[1]):
                    col = at[:, i]
                    out[f"ch{i+1}"] = {'sampling_rate': default_fs, 'data': _clean_channel_array(col)}
            return out
        else:
            # higher-dim fallback: flatten to 1d per channel approach (rare)
            flat = a.reshape(-1)
            out = {}
            for i, el in enumerate(flat):
                out[f"ch{i+1}"] = {'sampling_rate': default_fs, 'data': _clean_channel_array(np.asarray(el))}
            return out
    # for lists / tuples
    if isinstance(arr, (list, tuple)):
        out = {}
        for i, el in enumerate(arr):
            out[f"ch{i+1}"] = {'sampling_rate': default_fs, 'data': _clean_channel_array(np.asarray(el))}
        return out
    raise ValueError("Cannot convert supplied object to dictionary channels.")

def _clean_channel_array(data):
    """Apply stitch fixes and return 1D float array."""
    arr = np.asarray(data)
    # if 2D with two identical columns -> take first
    if arr.ndim == 2:
        if arr.shape[1] == 2:
            try:
                if np.allclose(arr[:, 0], arr[:, 1]):
                    arr = arr[:, 0]
                else:
                    # pick column with largest std-dev
                    idx = int(np.argmax(np.std(arr, axis=0)))
                    arr = arr[:, idx]
            except Exception:
                arr = arr[:, 0]
        else:
            # choose column with max std
            try:
                idx = int(np.argmax(np.std(arr, axis=0)))
                arr = arr[:, idx]
            except Exception:
                arr = arr[:, 0]
    arr = np.asarray(arr).squeeze()
    if arr.ndim != 1:
        arr = arr.ravel()
    n = len(arr)
    # adjacent duplicated pairs: [a,a,b,b] -> [a,b]
    if n > 2 and n % 2 == 0:
        try:
            if np.allclose(arr[0::2], arr[1::2]):
                arr = arr[0::2]
            else:
                half = n // 2
                if np.allclose(arr[:half], arr[half:]):
                    arr = arr[:half]
        except Exception:
            pass
    return arr.astype(float)

# -------------------------
# Main GUI
# -------------------------
class PeakAnalysisAppV034:
    def __init__(self):
        self.version = "v0.3.4"
        self.root = tk.Tk()
        self.root.title(f"PeakAnalysis {self.version} (Prototype)")
        self.root.geometry("900x540")

        # application state
        self.file_path = None
        self.channels = {}  # dict 'chN' -> { 'sampling_rate', 'data', 'label' }
        self.channel_keys = []
        self.selected_channel_key = None
        self.selected_entry = None
        self.results = []  # saved peaks list
        self.mem_popup_enabled = True  # whether to show peak-saved popup
        self.task_enabled = False
        self.task_start = None
        self.task_end = None

        # build UI
        self._build_intro_frame()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------- Intro frame (1)
    def _clear_root(self):
        for w in self.root.winfo_children():
            w.destroy()

    def _build_intro_frame(self):
        self._clear_root()
        frm = tk.Frame(self.root, padx=12, pady=12)
        frm.pack(fill="both", expand=True)
        intro_text = (
            "Welcome to PeakAnalysis v0.3.4 (Prototype) 🧠\n\n"
            "This interactive GUI analyzes power spectra of a signal channel to extract the "
            "beta frequency peak (13–30 Hz) and its amplitude.\n\n"
            "You can navigate through the time block flexibly using the arrow functions and time inputs.\n\n"
            "Designed and written by: Jawad Karim\n"
            #"Reviewed & edited by: Bara Kharseh\n"
            "Supervision: Dr. William D. Hutchison (Hutchison Lab, University of Toronto)\n\n"
            "For more code: https://github.com/wdhutchison-neurophysiolab\n"
            "Contact: j.karim@mail.utoronto.ca\n\n"
            "NOTE: This code does not preprocess the channels. You must preprocess channels before analyzing here."
        )
        lbl = tk.Label(frm, text=intro_text, justify="left", wraplength=840, font=("Segoe UI", 10))
        lbl.pack(fill="both", expand=True)
        btn = tk.Button(frm, text="Continue to Setup", command=self._build_config_frame, font=("Segoe UI", 11), padx=8, pady=6)
        btn.pack(pady=8)

    # ---------- Config frame (2)
    def _build_config_frame(self):
        self._clear_root()
        frm = tk.Frame(self.root, padx=12, pady=12)
        frm.pack(fill="both", expand=True)

        left = tk.Frame(frm)
        left.pack(side="left", fill="y")

        tk.Label(left, text="Enter .SMR or .NPY file path:", anchor="w").grid(row=0, column=0, sticky="w")
        self.file_entry = tk.Entry(left, width=58)
        self.file_entry.grid(row=1, column=0, sticky="w")
        tk.Button(left, text="Browse...", command=self._browse_file).grid(row=1, column=1, padx=(8, 0))

        # FFT combo default 16 now
        tk.Label(left, text="FFT window (nfft preference):", anchor="w").grid(row=2, column=0, sticky="w", pady=(10, 0))
        pow2 = list_power_of_two(min_pow=4, max_pow=18)
        self.fft_combo = ttk.Combobox(left, values=pow2, state="readonly", width=18)
        default_val = 16 if 16 in pow2 else pow2[0]
        self.fft_combo.set(default_val)
        self.fft_combo.grid(row=3, column=0, sticky="w")

        # start / end times
        tk.Label(left, text="Initial start time (s):").grid(row=4, column=0, sticky="w", pady=(8, 0))
        self.start_input = tk.Entry(left, width=14)
        self.start_input.insert(0, "0.0")
        self.start_input.grid(row=5, column=0, sticky="w")

        tk.Label(left, text="Initial end time (s):").grid(row=4, column=1, sticky="w", pady=(8, 0))
        self.end_input = tk.Entry(left, width=14)
        self.end_input.insert(0, "50.0")
        self.end_input.grid(row=5, column=1, sticky="w")

        # channel combo
        tk.Label(left, text="Channel:", anchor="w").grid(row=6, column=0, sticky="w", pady=(12, 0))
        self.channel_combo = ttk.Combobox(left, values=[], state="readonly", width=46)
        self.channel_combo.grid(row=7, column=0, columnspan=2, sticky="w")
        self.channel_combo.set("(open a file to list channels)")

        tk.Button(left, text="Preview Channel", command=self._preview_selected_channel).grid(row=8, column=0, pady=(8, 0))
        tk.Button(left, text="Start Analysis", command=self._start_analysis).grid(row=8, column=1, pady=(8, 0))

        # Task checkbox (2i)
        self.task_var = tk.BooleanVar(value=False)
        tk.Checkbutton(left, text="Is there a task in this recording?", variable=self.task_var, command=self._toggle_task_inputs).grid(row=9, column=0, columnspan=2, sticky="w", pady=(12, 0))
        tk.Label(left, text="Task start (s):").grid(row=10, column=0, sticky="w")
        tk.Label(left, text="Task end (s):").grid(row=10, column=1, sticky="w")
        self.task_start_entry = tk.Entry(left, width=12)
        self.task_end_entry = tk.Entry(left, width=12)
        self.task_start_entry.grid(row=11, column=0, sticky="w")
        self.task_end_entry.grid(row=11, column=1, sticky="w")
        self.task_start_entry.insert(0, "0.0")
        self.task_end_entry.insert(0, "30.0")
        # hide initially
        self.task_start_entry.grid_remove()
        self.task_end_entry.grid_remove()

        # notes
        notes = (
            "Notes:\n"
            "- 'nfft preference' will be used as the FFT length (nfft >= nperseg).\n"
            "- This GUI does NOT preprocess channels. Preprocess before analysis.\n"
            "- We recommend single-channel files for analysis to avoid selection/format issues.\n"
            "- If using .npy, it should ideally be a dictionary mapping channel names to {'sampling_rate','data'}.\n"
            "  Example: {'ch1':{'sampling_rate':125.0,'data':np.array([...])}, ...}\n"
            "- If you see a load error about Python scalar conversion, you can let PeakAnalysis convert the .npy to dictionary."
        )
        tk.Label(left, text=notes, wraplength=360, justify="left", fg="gray").grid(row=12, column=0, columnspan=2, pady=(10, 0))

        # right column: status & memory preview
        right = tk.Frame(frm, padx=12)
        right.pack(side="left", fill="both", expand=True)

        self.status_label = tk.Label(right, text="No file loaded", anchor="w", justify="left")
        self.status_label.pack(fill="x", pady=(6, 0))

        tk.Label(right, text="Saved Peaks (in memory):", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.mem_preview = tk.Listbox(right, height=12)
        self.mem_preview.pack(fill="both", expand=True, padx=(0, 4))
        tk.Button(right, text="Open Memory Window", command=self._open_memory_window).pack(pady=(6, 0))

    def _toggle_task_inputs(self):
        if self.task_var.get():
            self.task_start_entry.grid()
            self.task_end_entry.grid()
        else:
            self.task_start_entry.grid_remove()
            self.task_end_entry.grid_remove()

    # ---------- file browsing and conversion (2a, 2b)
    def _browse_file(self):
        fp = filedialog.askopenfilename(title="Select .SMR or .NPY file", filetypes=[("SMR files","*.smr"),("NPY files","*.npy"),("All files","*.*")])
        if not fp:
            return
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, fp)
        self.file_path = fp
        try:
            if fp.lower().endswith('.smr'):
                if neo is None:
                    raise RuntimeError("neo package not available - cannot read SMR files.")
                reader = neo.io.Spike2IO(filename=fp)
                block = reader.read_block()
                seg = block.segments[0]
                analogs = seg.analogsignals
                if len(analogs) == 0:
                    raise ValueError("No analog signals found in the SMR file.")
                self.channels = {}
                for idx, sig in enumerate(analogs):
                    try:
                        sr = get_sampling_rate_hz(sig)
                    except Exception:
                        sr = None
                    data = np.asarray(sig).squeeze()
                    data = _clean_channel_array(data)
                    label = getattr(sig, 'name', None)
                    if label is None:
                        ann = getattr(sig, 'annotations', None)
                        if isinstance(ann, dict):
                            label = ann.get('name') or ann.get('channel_name') or ann.get('label')
                    key = f"ch{idx+1}"
                    self.channels[key] = {'sampling_rate': float(sr) if sr is not None else None, 'data': data, 'label': label or key}
                self._populate_channels()
                self.status_label.config(text=f"Loaded SMR: {os.path.basename(fp)} | Channels: {len(self.channels)}")
            elif fp.lower().endswith('.npy'):
                # inspect
                is_dict, info, loaded = inspect_npy_file(fp)
                if is_dict:
                    # already dict
                    # loaded may be dict object or raw dict
                    if isinstance(loaded, dict):
                        # convert entries to normalized format
                        self.channels = {}
                        idx = 1
                        for k, v in loaded.items():
                            if isinstance(v, dict) and 'data' in v:
                                data = _clean_channel_array(v['data'])
                                sr = v.get('sampling_rate', None)
                            else:
                                data = _clean_channel_array(v)
                                sr = None
                            self.channels[f"ch{idx}"] = {'sampling_rate': float(sr) if sr is not None else None, 'data': data, 'label': str(k)}
                            idx += 1
                        self._populate_channels()
                        self.status_label.config(text=f"Loaded NPY(dict): {os.path.basename(fp)} | Channels: {len(self.channels)}")
                    else:
                        # fallback
                        self.channels = force_convert_npy_to_dict(loaded)
                        self._populate_channels()
                        self.status_label.config(text=f"Loaded NPY(dict-like): {os.path.basename(fp)} | Channels: {len(self.channels)}")
                else:
                    # not dict — offer to convert
                    ans = messagebox.askyesno("NPY not dictionary", f"The .npy appears to be:\n{info}\n\nPeakAnalysis {self.version} can attempt to convert this .npy to dictionary format (ch1/ch2/...). Do you want to convert it now?")
                    if not ans:
                        # load raw as array and wrap as ch1
                        arr = loaded
                        try:
                            converted = force_convert_npy_to_dict(arr)
                        except Exception as exc:
                            messagebox.showerror("Convert failed", f"Auto-conversion failed:\n{exc}")
                            return
                        self.channels = converted
                        self._populate_channels()
                        self.status_label.config(text=f"Loaded NPY (auto-wrapped, not saved): {os.path.basename(fp)} | Channels: {len(self.channels)}")
                        return
                    # convert
                    try:
                        arr = loaded
                        converted = force_convert_npy_to_dict(arr)
                        # converted is in memory; offer to save dictionary to disk
                        save_choice = messagebox.askyesno("Save converted?", "File converted to dictionary successfully. Would you like to save the dictionary-format file to disk? (Yes -> choose path; No -> use in memory)")
                        if save_choice:
                            save_fp = filedialog.asksaveasfilename(defaultextension='.npy', filetypes=[('NPY files','*.npy'),('All files','*.*')])
                            if save_fp:
                                np.save(save_fp, converted, allow_pickle=True)
                                messagebox.showinfo("Saved", f"Saved converted dictionary to:\n{save_fp}")
                        self.channels = converted
                        self._populate_channels()
                        self.status_label.config(text=f"Loaded NPY(converted): {os.path.basename(fp)} | Channels: {len(self.channels)}")
                    except Exception as exc:
                        messagebox.showerror("Conversion failed", f"Failed converting .npy to dictionary:\n{exc}")
                        return
            else:
                messagebox.showerror("File type", "Unsupported file type. Please choose a .smr or .npy file.")
                return
            # after loading channels, if any missing sampling rates ask user to enter
            need_sr = False
            for k, v in self.channels.items():
                if v.get('sampling_rate') is None:
                    need_sr = True
                    break
            if need_sr:
                # ask user if they want to input sampling rate to apply to all
                if messagebox.askyesno("Sampling rate missing", "One or more channels do not have a sampling rate. Would you like to enter a sampling rate to apply to all channels?"):
                    val = simpledialog.askfloat("Sampling rate", "Enter sampling rate (Hz)", minvalue=1.0)
                    if val is not None:
                        for k in self.channels:
                            if self.channels[k].get('sampling_rate') is None:
                                self.channels[k]['sampling_rate'] = float(val)
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Load error", f"Failed to load file:\n{exc}")
            self.status_label.config(text="No file loaded (error)")

    def _populate_channels(self):
        self.channel_keys = list(self.channels.keys())
        choices = []
        for i, k in enumerate(self.channel_keys):
            label = self.channels[k].get('label') or k
            choices.append(f"channel {i+1} ({k}) - {label}")
        self.channel_combo.config(values=choices)
        if choices:
            self.channel_combo.set(choices[0])
            self.selected_channel_key = self.channel_keys[0]
            self.selected_entry = self.channels[self.selected_channel_key]

    def _get_selected_channel_key(self):
        val = self.channel_combo.get()
        if not val:
            return None
        try:
            part = val.split('(')[1]
            key = part.split(')')[0].strip()
            return key
        except Exception:
            return self.channel_keys[0] if self.channel_keys else None

    # ---------- Preview (entire signal) (2f)
    def _preview_selected_channel(self):
        if not self.channel_keys:
            messagebox.showinfo("No file", "Please open a file with channels first.")
            return
        key = self._get_selected_channel_key()
        if key is None or key not in self.channels:
            messagebox.showerror("Channel error", "Selected channel key not found.")
            return
        entry = self.channels[key]
        data = np.asarray(entry['data'])
        sr = entry.get('sampling_rate', None)
        if sr is None:
            sr_text = "Unknown"
            rec_len = float(len(data))
        else:
            sr_text = f"{sr:.2f} Hz"
            rec_len = float(len(data)) / float(sr)
        # handle 2D duplicates
        if data.ndim > 1:
            if data.shape[1] == 2 and np.allclose(data[:, 0], data[:, 1]):
                data = data[:, 0]
            else:
                if data.shape[1] > 1:
                    if not messagebox.askyesno("Multi-column data", "Channel data has multiple columns. Use first column for preview?"):
                        # user canceled; still show first column but warn
                        pass
                    data = data[:, 0]
        win = tk.Toplevel(self.root)
        win.title(f"Preview - {key}")
        win.geometry("1000x380")
        fig = Figure(figsize=(9.5, 3.5), dpi=100)
        ax = fig.add_subplot(111)
        t = np.arange(len(data)) / float(sr) if sr is not None else np.arange(len(data))
        ax.plot(t, data, color='black', linewidth=0.6)
        ax.set_xlabel('Time (s)')
        ax.set_title(f"{key} | Sampling rate: {sr_text} | Duration: {rec_len:.2f} s | Samples: {len(data)}")
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        # bottom controls
        frm = tk.Frame(win)
        frm.pack(pady=6)
        def on_confirm():
            self.selected_channel_key = key
            self.selected_entry = {'data': data, 'sampling_rate': sr, 'label': entry.get('label', key)}
            win.destroy()
            try:
                nfft_val = int(self.fft_combo.get())
            except Exception:
                nfft_val = int(self.fft_combo['values'][0])
            try:
                s0 = float(self.start_input.get())
                e0 = float(self.end_input.get())
            except Exception:
                s0, e0 = 0.0, min(50.0, rec_len)
            self._launch_spectrum_explorer(key, nfft_val, s0, e0)
        tk.Button(frm, text="Confirm & Open Explorer", command=on_confirm).pack(side='left', padx=8)
        tk.Button(frm, text="Close", command=win.destroy).pack(side='left', padx=8)

    # ---------- Start analysis -> launch explorer (3)
    def _start_analysis(self):
        key = self._get_selected_channel_key()
        if not key:
            messagebox.showerror("No channel", "Select a channel first.")
            return
        entry = self.channels.get(key)
        if entry is None:
            messagebox.showerror("Channel error", "Selected channel missing.")
            return
        self.selected_channel_key = key
        self.selected_entry = entry
        try:
            nfft_val = int(self.fft_combo.get())
        except Exception:
            nfft_val = int(self.fft_combo['values'][0])
        try:
            s0 = float(self.start_input.get())
            e0 = float(self.end_input.get())
        except Exception:
            sr = entry.get('sampling_rate')
            rec_len = float(len(entry['data'])) / float(sr) if sr else float(len(entry['data']))
            s0, e0 = 0.0, min(50.0, rec_len)
        # task info
        self.task_enabled = bool(self.task_var.get())
        if self.task_enabled:
            try:
                self.task_start = float(self.task_start_entry.get())
                self.task_end = float(self.task_end_entry.get())
            except Exception:
                messagebox.showerror("Task times", "Enter valid numeric task start/end times.")
                return
        self._launch_spectrum_explorer(self.selected_channel_key, nfft_val, s0, e0)

    def _launch_spectrum_explorer(self, channel_key, nfft_pref, start_time_init, end_time_init):
        entry = self.channels[channel_key]
        sr = entry.get('sampling_rate', None)
        data = np.asarray(entry['data']).astype(float)
        if sr is None:
            if messagebox.askyesno("Sampling rate missing", "Sampling rate not found. Enter manually?"):
                val = simpledialog.askfloat("Sampling rate", "Enter sampling rate in Hz", minvalue=1.0)
                if val is None:
                    messagebox.showerror("No sampling rate", "Sampling rate required for spectral analysis.")
                    return
                sr = float(val)
            else:
                messagebox.showerror("No sampling rate", "Sampling rate required for spectral analysis.")
                return
        # clean duplicated/stitched issues
        if data.ndim == 1 and len(data) % 2 == 0:
            try:
                if np.allclose(data[0::2], data[1::2]):
                    data = data[::2]
            except Exception:
                pass
        if data.ndim == 2 and data.shape[1] == 2:
            try:
                if np.allclose(data[:, 0], data[:, 1]):
                    data = data[:, 0]
            except Exception:
                pass
        rec_length_s = float(len(data)) / float(sr)
        state = {
            'current_start': float(max(0.0, start_time_init)),
            'current_end': float(min(end_time_init, rec_length_s)),
            'nfft_pref': int(nfft_pref),
            'sampling_rate': float(sr),
            'lfp': data,
            'rec_len': rec_length_s,
            'channel_key': channel_key,
            'last_result': None
        }

        # Explorer window
        win = tk.Toplevel(self.root)
        win.title(f"Spectrum Explorer - {channel_key}")
        win.geometry("980x680")

        fig = Figure(figsize=(9.5, 4.8), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_xlabel('Frequency (Hz)')
        ax.set_ylabel('Power per bin (V^2)')
        ax.grid(True, linestyle='--', alpha=0.5)
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(side='top', fill='both', expand=True)
        toolbar_frame = tk.Frame(win)
        toolbar_frame.pack(side='top', fill='x')
        toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
        toolbar.update()

        ctrl = tk.Frame(win, padx=8, pady=6)
        ctrl.pack(fill='x')
        fn_label_text = f"File: {os.path.basename(self.file_path) if self.file_path else 'N/A'}  Recording: {rec_length_s:.1f}s  Fs: {sr:.1f} Hz  nfft(pref): {state['nfft_pref']}"
        tk.Label(ctrl, text=fn_label_text).grid(row=0, column=0, columnspan=9, sticky='w')

        tk.Label(ctrl, text='Start (s):').grid(row=1, column=0, sticky='e')
        start_var = tk.StringVar(value=f"{state['current_start']:.3f}")
        start_ent = tk.Entry(ctrl, textvariable=start_var, width=10)
        start_ent.grid(row=1, column=1, sticky='w')

        tk.Label(ctrl, text='End (s):').grid(row=1, column=2, sticky='e')
        end_var = tk.StringVar(value=f"{state['current_end']:.3f}")
        end_ent = tk.Entry(ctrl, textvariable=end_var, width=10)
        end_ent.grid(row=1, column=3, sticky='w')

        def compute_and_plot():
            try:
                s = float(start_var.get()); e = float(end_var.get())
            except Exception:
                messagebox.showerror('Time error', 'Enter valid numeric start/end times.')
                return
            if s < 0 or e <= s or s >= state['rec_len']:
                messagebox.showerror('Time error', 'Invalid start/end times for this recording.')
                return
            sr_local = state['sampling_rate']
            start_idx = int(round(s * sr_local)); end_idx = int(round(e * sr_local))
            if end_idx <= start_idx + 1:
                messagebox.showwarning('Window too short', 'Selected window is too short for spectral estimation.')
                return
            seg_data = state['lfp'][start_idx:end_idx]
            preferred = state['nfft_pref']
            nperseg = min(preferred, len(seg_data))
            if nperseg < 8:
                nperseg = len(seg_data)
            nfft_used = max(preferred, nperseg)
            freqs, pxx = welch(seg_data, fs=sr_local, nperseg=nperseg, nfft=nfft_used, scaling='density')
            if len(freqs) < 2:
                messagebox.showwarning('Spectrum', 'Could not compute spectrum (too few freq bins).')
                return
            df = freqs[1] - freqs[0]
            pxx_bin = pxx * df
            peak_freq, peak_power = find_beta_peak(freqs, pxx_bin, beta_band=(13.0, 30.0))

            ax.clear()
            ax.bar(freqs, pxx_bin, width=df, align='center', edgecolor='white', linewidth=0.6, color='black')
            if not math.isnan(peak_freq):
                ax.axvline(peak_freq, color='red', linestyle='--', linewidth=1.2)
                ax.scatter([peak_freq], [peak_power], color='red', zorder=20)
                annotation = f"Peak: {peak_freq:.2f} Hz\nPower: {peak_power:.9f}"
            else:
                annotation = "No beta band data"
            ax.text(0.98, 0.95, annotation, transform=ax.transAxes, fontsize=12, color='red', ha='right', va='top', bbox=dict(facecolor='white', alpha=0.8, edgecolor='red'))
            nyq = 0.5 * sr_local
            ax.set_xlim(0, min(50.0, nyq))
            ax.set_xlabel('Frequency (Hz)')
            ax.set_ylabel('Power per bin (V^2)')
            ax.set_title(f"Welch Spectrum ({s:.3f}–{e:.3f} s) | Channel: {state['channel_key']}")
            ax.grid(True, linestyle='--', alpha=0.4)
            canvas.draw()
            # compute task overlap if any
            is_task_block = False
            overlap_type = 'none'
            if self.task_var.get():
                try:
                    ts = float(self.task_start_entry.get()); te = float(self.task_end_entry.get())
                    # determine overlap
                    if s >= ts and e <= te:
                        is_task_block = True
                        overlap_type = 'full'
                    elif (s < te and e > ts):
                        # partial overlap
                        is_task_block = True
                        overlap_type = 'partial'
                    else:
                        is_task_block = False
                        overlap_type = 'none'
                except Exception:
                    is_task_block = False
                    overlap_type = 'none'
            # store last result
            state['last_result'] = {
                'file': os.path.basename(self.file_path) if self.file_path else '',
                'channel': state['channel_key'],
                'start_s': s,
                'end_s': e,
                'peak_freq_hz': peak_freq,
                'peak_power': peak_power,
                'nperseg': nperseg,
                'nfft': nfft_used,
                'is_task': bool(is_task_block),
                'task_overlap': overlap_type
            }

        btn_proc = tk.Button(ctrl, text='Process', command=compute_and_plot)
        btn_proc.grid(row=1, column=4, padx=(8,4))

        # 1s shift
        def shift_left_1s():
            try:
                s = float(start_var.get()); e = float(end_var.get())
            except Exception:
                return
            s_new = max(0.0, s - 1.0); e_new = max(s_new + 0.001, e - 1.0)
            start_var.set(f"{s_new:.3f}"); end_var.set(f"{e_new:.3f}"); compute_and_plot()
        def shift_right_1s():
            try:
                s = float(start_var.get()); e = float(end_var.get())
            except Exception:
                return
            if e + 1.0 <= state['rec_len']:
                s_new = s + 1.0; e_new = e + 1.0
            else:
                s_new = min(s + 1.0, state['rec_len'] - 0.001); e_new = state['rec_len']
            start_var.set(f"{s_new:.3f}"); end_var.set(f"{e_new:.3f}"); compute_and_plot()
        tk.Button(ctrl, text='◀ 1s', command=shift_left_1s).grid(row=2, column=0, pady=(6,0))
        tk.Button(ctrl, text='1s ▶', command=shift_right_1s).grid(row=2, column=1, pady=(6,0))

        # fine step
        tk.Label(ctrl, text='Step (s)').grid(row=2, column=2)
        step_var = tk.StringVar(value='0.1')
        step_ent = tk.Entry(ctrl, textvariable=step_var, width=8)
        step_ent.grid(row=2, column=3)
        def shift_left_fine():
            try:
                step = float(step_var.get()); s = float(start_var.get()); e = float(end_var.get())
            except Exception:
                return
            s_new = max(0.0, s - step); e_new = max(s_new + 0.001, e - step)
            start_var.set(f"{s_new:.3f}"); end_var.set(f"{e_new:.3f}"); compute_and_plot()
        def shift_right_fine():
            try:
                step = float(step_var.get()); s = float(start_var.get()); e = float(end_var.get())
            except Exception:
                return
            if e + step <= state['rec_len']:
                s_new = s + step; e_new = e + step
            else:
                s_new = min(s + step, state['rec_len'] - 0.001); e_new = state['rec_len']
            start_var.set(f"{s_new:.3f}"); end_var.set(f"{e_new:.3f}"); compute_and_plot()
        tk.Button(ctrl, text='◀ Fine', command=shift_left_fine).grid(row=2, column=4)
        tk.Button(ctrl, text='Fine ▶', command=shift_right_fine).grid(row=2, column=5)

        # Save last result to memory (3c) with popup and "Don't show again"
        def save_last_result():
            res = state.get('last_result')
            if not res:
                messagebox.showinfo('No result', 'No computed peak to save. Press Process first.')
                return
            self.results.append(res)
            summary = f"{res['file']} | {res['channel']} | {res['start_s']:.2f}-{res['end_s']:.2f}s | {res['peak_freq_hz']:.2f} Hz"
            self.mem_preview.insert(tk.END, summary)
            # show popup with option "Don't show this again"
            if self.mem_popup_enabled:
                popup = tk.Toplevel(win)
                popup.title("Peak saved")
                tk.Label(popup, text=f"Peak saved: {res['peak_freq_hz']:.3f} Hz\nPower: {res['peak_power']:.9f}").pack(padx=12, pady=8)
                dont_var = tk.BooleanVar(value=False)
                tk.Checkbutton(popup, text="Don't show this again", variable=dont_var).pack(pady=(0,8))
                def ok_close():
                    if dont_var.get():
                        self.mem_popup_enabled = False
                    popup.destroy()
                tk.Button(popup, text="OK", command=ok_close).pack(pady=(0,8))
            else:
                # silent append
                pass

        tk.Button(ctrl, text='Save peak (to memory)', command=save_last_result).grid(row=1, column=5, padx=(8,4))

        # Export CSV (new file) (3d)
        def export_csv():
            if not self.results:
                messagebox.showinfo('No data', 'No results to export.')
                return
            fp = filedialog.asksaveasfilename(defaultextension='.csv', filetypes=[('CSV files','*.csv'),('All files','*.*')])
            if not fp: return
            try:
                with open(fp, 'w', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap'])
                    writer.writeheader()
                    for r in self.results:
                        writer.writerow(r)
                messagebox.showinfo('Exported', f'Exported {len(self.results)} rows to {fp}')
            except Exception as exc:
                messagebox.showerror('Export error', f'Failed saving CSV:\n{exc}')
        tk.Button(ctrl, text='Export CSV', command=export_csv).grid(row=1, column=6, padx=(8,4))

        # Export into existing spreadsheet (3e) - complex options
        def export_into_existing():
            if not self.results:
                messagebox.showinfo('No data', 'No results to export.')
                return
            fp = filedialog.askopenfilename(title="Select existing CSV or XLSX file", filetypes=[('CSV files','*.csv'),('Excel files','*.xlsx;*.xls'),('All files','*.*')])
            if not fp: return
            # detect type
            is_csv = fp.lower().endswith('.csv')
            is_xlsx = fp.lower().endswith('.xlsx') or fp.lower().endswith('.xls')
            if is_csv:
                # no sheets, append directly
                try:
                    file_exists = os.path.exists(fp)
                    with open(fp, 'a', newline='') as f:
                        writer = csv.DictWriter(f, fieldnames=['file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap'])
                        if not file_exists or os.path.getsize(fp) == 0:
                            writer.writeheader()
                        for r in self.results:
                            writer.writerow(r)
                    messagebox.showinfo('Exported', f'Appended {len(self.results)} rows to {fp}')
                except PermissionError:
                    messagebox.showerror('Permission error', 'The file is probably open. Please close it and try again.')
                except Exception as exc:
                    messagebox.showerror('Export error', f'Failed exporting to CSV:\n{exc}')
            elif is_xlsx:
                if openpyxl is None:
                    messagebox.showerror('openpyxl missing', 'openpyxl required to manipulate .xlsx files. Install it or export to CSV.')
                    return
                try:
                    wb = load_workbook(fp)
                except PermissionError:
                    messagebox.showerror('Permission error', 'The file is probably open. Please close it and try again.')
                    return
                except Exception as exc:
                    messagebox.showerror('Workbook error', f'Failed opening workbook: {exc}')
                    return
                # build sheet list and ask user what to do
                sheets = wb.sheetnames
                # popup to choose options
                opt_win = tk.Toplevel(win)
                opt_win.title("Export options into workbook")
                tk.Label(opt_win, text="Choose export mode:").pack(anchor='w', padx=8, pady=(8,4))
                mode_var = tk.StringVar(value='all_in_selected')
                tk.Radiobutton(opt_win, text="All rows into a selected existing sheet (or create new)", variable=mode_var, value='all_in_selected').pack(anchor='w', padx=8)
                tk.Radiobutton(opt_win, text="Each file into a separate sheet in this workbook", variable=mode_var, value='each_in_sheet').pack(anchor='w', padx=8)
                tk.Radiobutton(opt_win, text="Create separate new files for each analyzed file (CSV/XLSX per file)", variable=mode_var, value='separate_files').pack(anchor='w', padx=8)
                tk.Label(opt_win, text="Select sheet (for modes that use existing sheets):").pack(anchor='w', padx=8, pady=(8,0))
                sheet_combo = ttk.Combobox(opt_win, values=sheets + ['<create new>'], state='readonly', width=40)
                if sheets:
                    sheet_combo.set(sheets[0])
                else:
                    sheet_combo.set('<create new>')
                sheet_combo.pack(anchor='w', padx=8, pady=(0,8))
                new_sheet_name_var = tk.StringVar(value=os.path.splitext(os.path.basename(self.file_path))[0] if self.file_path else 'new_sheet')
                tk.Label(opt_win, text="If creating new sheet, enter new sheet name:").pack(anchor='w', padx=8)
                new_sheet_entry = tk.Entry(opt_win, textvariable=new_sheet_name_var, width=36)
                new_sheet_entry.pack(anchor='w', padx=8, pady=(0,8))
                # submit
                def do_export_choice():
                    mode = mode_var.get()
                    selected_sheet = sheet_combo.get()
                    new_sheet_name = new_sheet_name_var.get()[:31]
                    if mode == 'all_in_selected':
                        # determine sheet to use
                        target_sheet = None
                        if selected_sheet == '<create new>':
                            # create new
                            target_sheet = new_sheet_name
                            if target_sheet in wb.sheetnames:
                                i = 1
                                while f"{target_sheet}_{i}" in wb.sheetnames:
                                    i += 1
                                target_sheet = f"{target_sheet}_{i}"
                            ws = wb.create_sheet(title=target_sheet)
                        else:
                            target_sheet = selected_sheet
                            ws = wb[target_sheet]
                        # append header if empty
                        headers = ['file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap']
                        # find first empty row
                        row0 = ws.max_row
                        if row0 == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                            ws.append(headers)
                        for r in self.results:
                            ws.append([r.get(h) for h in headers])
                        try:
                            wb.save(fp)
                            messagebox.showinfo('Exported', f'Appended {len(self.results)} rows to sheet "{target_sheet}" in {fp}')
                        except PermissionError:
                            messagebox.showerror('Permission error', 'File likely open. Close it and try again.')
                        except Exception as exc:
                            messagebox.showerror('Save error', f'Failed saving workbook:\n{exc}')
                    elif mode == 'each_in_sheet':
                        # for each file in results find file name and group rows
                        grouped = {}
                        for r in self.results:
                            fname = r.get('file') or 'unknown'
                            grouped.setdefault(fname, []).append(r)
                        for fname, rows in grouped.items():
                            sheet_name = os.path.splitext(fname)[0][:31]
                            if sheet_name in wb.sheetnames:
                                i = 1
                                base = sheet_name
                                while f"{base}_{i}" in wb.sheetnames:
                                    i += 1
                                sheet_name = f"{base}_{i}"
                            ws = wb.create_sheet(title=sheet_name)
                            headers = ['file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap']
                            ws.append(headers)
                            for r in rows:
                                ws.append([r.get(h) for h in headers])
                        try:
                            wb.save(fp)
                            messagebox.showinfo('Exported', f'Wrote {len(grouped)} new sheets to {fp}')
                        except PermissionError:
                            messagebox.showerror('Permission error', 'File likely open. Close it and try again.')
                        except Exception as exc:
                            messagebox.showerror('Save error', f'Failed saving workbook:\n{exc}')
                    elif mode == 'separate_files':
                        # ask user for folder to write into
                        folder = filedialog.askdirectory(title="Select folder to write individual files")
                        if not folder:
                            return
                        grouped = {}
                        for r in self.results:
                            fname = r.get('file') or 'unknown'
                            grouped.setdefault(fname, []).append(r)
                        for fname, rows in grouped.items():
                            base = os.path.splitext(fname)[0]
                            out_csv = os.path.join(folder, f"{base}.csv")
                            with open(out_csv, 'w', newline='') as f:
                                writer = csv.DictWriter(f, fieldnames=['file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap'])
                                writer.writeheader()
                                for rr in rows:
                                    writer.writerow(rr)
                        messagebox.showinfo('Exported', f'Wrote {len(grouped)} CSV files to {folder}')
                    opt_win.destroy()
                tk.Button(opt_win, text="Export", command=do_export_choice).pack(pady=(8,8))
            else:
                messagebox.showinfo('Unsupported', 'Only .csv and .xlsx are supported for this operation.')

        tk.Button(ctrl, text='Export into existing spreadsheet', command=export_into_existing).grid(row=1, column=7, padx=(8,4))

        # initialize and compute
        start_var.set(f"{state['current_start']:.3f}")
        end_var.set(f"{state['current_end']:.3f}")
        compute_and_plot()

    # ---------- Memory window (3c)
    def _open_memory_window(self):
        win = tk.Toplevel(self.root)
        win.title("Saved Peaks (Memory)")
        win.geometry("760x360")
        frame = tk.Frame(win, padx=8, pady=8)
        frame.pack(fill='both', expand=True)
        cols = ('file','channel','start_s','end_s','peak_freq_hz','peak_power','nperseg','nfft','is_task','task_overlap')
        tree = ttk.Treeview(frame, columns=cols, show='headings', height=12)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=80, anchor='center')
        tree.pack(fill='both', expand=True)
        for r in self.results:
            tree.insert('', tk.END, values=(r.get('file',''), r.get('channel',''), f"{r.get('start_s',0):.3f}", f"{r.get('end_s',0):.3f}", f"{r.get('peak_freq_hz',float('nan')):.3f}", f"{r.get('peak_power',0):.9f}", r.get('nperseg'), r.get('nfft'), r.get('is_task',False), r.get('task_overlap','none')))
        btn_frame = tk.Frame(win)
        btn_frame.pack(fill='x', pady=(8,0))
        def export_csv_local():
            if not self.results:
                messagebox.showinfo('No data', 'No results to export.')
                return
            fp = filedialog.asksaveasfilename(defaultextension='.csv', filetypes=[('CSV files','*.csv'),('All files','*.*')])
            if not fp: return
            try:
                with open(fp, 'w', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=cols)
                    writer.writeheader()
                    for r in self.results:
                        writer.writerow(r)
                messagebox.showinfo('Exported', f'Exported {len(self.results)} rows to {fp}')
            except Exception as exc:
                messagebox.showerror('Export error', f'Failed saving CSV:\n{exc}')
        tk.Button(btn_frame, text="Export saved peaks to CSV", command=export_csv_local).pack(side='left', padx=4)
        tk.Button(btn_frame, text="Close", command=win.destroy).pack(side='right', padx=4)

    # ---------- shutdown
    def _on_close(self):
        if messagebox.askokcancel("Quit", "Do you want to quit PeakAnalysis?"):
            self.root.destroy()

    def run(self):
        self.root.mainloop()

# -------------------------
# Entrypoint
# -------------------------
if __name__ == '__main__':
    app = PeakAnalysisAppV034()
    app.run()
